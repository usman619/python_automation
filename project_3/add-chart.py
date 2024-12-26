from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference

wb = load_workbook('pivot_table.xlsx')

sheet = wb['Report']

min_col = wb.active.min_column
max_col = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row

bar_chart = BarChart()

data = Reference(sheet, min_col=min_col+1, max_col=max_col, min_row=min_row, max_row=max_row)
categories = Reference(sheet, min_col=min_col, max_col=min_col, min_row=min_row+1, max_row=max_row)

bar_chart.add_data(data=data, titles_from_data=True)
bar_chart.set_categories(categories)

sheet.add_chart(bar_chart, "B12")

bar_chart.title= "Sales by Product Line"
bar_chart.style = 5

wb.save('barchart.xlsx')
