from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

import os
import sys

from datetime import datetime

# importing the table 
application_path = os.path.dirname(sys.executable)
input_path = os.path.join(application_path,'pivot_table.xlsx')

wb = load_workbook(input_path)
sheet = wb['Report']

# getting current month in letters
current_month = datetime.now().strftime('%B').lower()
# current_month = input("Current Month: ")

min_col = wb.active.min_column
max_col = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row

# Applying formula on the pivot table
for i in range(min_col+1, max_col+1):
    letter = get_column_letter(i)
    sheet[f'{letter}{max_row+1}'] = f'=SUM({letter}{min_row+1}:{letter}{max_row})'
    sheet[f'{letter}{max_row+1}'].style = 'Currency'

# Making Barchart
bar_chart = BarChart()

data = Reference(sheet, min_col=min_col+1, max_col=max_col, min_row=min_row, max_row=max_row)
categories = Reference(sheet, min_col=min_col, max_col=min_col, min_row=min_row+1, max_row=max_row)

bar_chart.add_data(data=data, titles_from_data=True)
bar_chart.set_categories(categories)

sheet.add_chart(bar_chart, "B12")

bar_chart.title= "Sales by Product Line"
bar_chart.style = 5

# Adding title and month on the top of the sheet
sheet['A1'] = 'Sales Report'
sheet['A2'] = current_month

sheet['A1'].font = Font('Arial', bold=True, size=20)
sheet['A2'].font = Font('Arial', bold=True, size=10)

# Saving the file
file_name = f'report_{current_month}.xlsx'
output_path = os.path.join(application_path, file_name)

wb.save(output_path)
