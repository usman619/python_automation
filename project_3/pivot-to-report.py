from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

from datetime import datetime

wb = load_workbook('pivot_table.xlsx')
sheet = wb['Report']

current_month = datetime.now().strftime('%B').lower()

min_col = wb.active.min_column
max_col = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row

for i in range(min_col+1, max_col+1):
    letter = get_column_letter(i)
    sheet[f'{letter}{max_row+1}'] = f'=SUM({letter}{min_row+1}:{letter}{max_row})'
    sheet[f'{letter}{max_row+1}'].style = 'Currency'


bar_chart = BarChart()

data = Reference(sheet, min_col=min_col+1, max_col=max_col, min_row=min_row, max_row=max_row)
categories = Reference(sheet, min_col=min_col, max_col=min_col, min_row=min_row+1, max_row=max_row)

bar_chart.add_data(data=data, titles_from_data=True)
bar_chart.set_categories(categories)

sheet.add_chart(bar_chart, "B12")

bar_chart.title= "Sales by Product Line"
bar_chart.style = 5

sheet['A1'] = 'Sales Report'
sheet['A2'] = current_month

sheet['A1'].font = Font('Arial', bold=True, size=20)
sheet['A2'].font = Font('Arial', bold=True, size=10)

wb.save(f'/home/usman619/Desktop/report_{current_month}.xlsx')
